import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename): #Defining a function.
    wb = xl.load_workbook(filename) #Importing the whole workbook in wb.
    sheet = wb['Sheet1']                      #the items in workbook are sheets. we are selecting 'Sheet1'.
    for row in range(2, sheet.max_row + 1):   #Iterating through range. range is starting from 2nd row till last row + 1.
        cell = sheet.cell(row, 3)             #selecting different cells. row = 2,3,4. column = 3. 
        corrected_price = cell.value * 0.9    #making changes in the cells.
        corrected_price_cell = sheet.cell(row, 4)        #selecting row (2,3,4) in column 4.
        corrected_price_cell.value = corrected_price     #assigning corrected_price to corrected_price_cell
    values = Reference(sheet,                 #creating a reference in a new variable values.
                       min_row=2,      
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)
    chart = BarChart()
    chart.add_data(values)                    #adding values to chart
    sheet.add_chart(chart, 'e2')              #adding chart to sheet
    wb.save(filename)                         #saving the workbook

process_workbook('transactions.xlsx')         #using function for the specific file, 'transactions.xlsx'